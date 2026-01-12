# app/submissions.py
from __future__ import annotations

import hmac
import json
import os
import re
import unicodedata
from hashlib import sha256
from pathlib import Path
from typing import Dict, List, Any

from fastapi import APIRouter, HTTPException, Request, Form
from fastapi.responses import HTMLResponse, RedirectResponse
from fastapi.templating import Jinja2Templates
from urllib.parse import quote_plus, unquote_plus


# =================
# Config & commons
# =================
BASE_DIR = Path(__file__).resolve().parent        # app/
ROOT_DIR = BASE_DIR.parent                        # raíz del proyecto
TEMPLATES = Jinja2Templates(directory=str(BASE_DIR / "templates"))

SUBMIT_SECRET = os.getenv("SUBMIT_SECRET", "devsecret")
ADMIN_TOKEN = os.getenv("ADMIN_TOKEN", "devadmin")

PENDING_FILE = BASE_DIR / "pending_submissions.json"
APPROVED_FILE = BASE_DIR / "approved_submissions.json"

FIELD_ORDER = ("mid", "season", "date", "division", "j1", "j2")


def _canonical_query(params: Dict[str, str]) -> str:
    from urllib.parse import quote_plus as _qp
    parts = []
    for k in FIELD_ORDER:
        parts.append(f"{k}={_qp(params[k])}")
    return "&".join(parts)


def _make_sig(params: Dict[str, str]) -> str:
    canon = _canonical_query(params).encode("utf-8")
    return hmac.new(SUBMIT_SECRET.encode("utf-8"), canon, sha256).hexdigest()


def _verify_sig(params: Dict[str, str], sig: str) -> None:
    expected = _make_sig(params)
    if not hmac.compare_digest(expected, sig):
        raise HTTPException(status_code=400, detail="Enlace inválido o manipulado")


def _load_json(path: Path) -> List[Dict[str, Any]]:
    if path.exists():
        try:
            return json.loads(path.read_text(encoding="utf-8"))
        except Exception:
            return []
    return []


def _save_json(path: Path, data: List[Dict[str, Any]]) -> None:
    path.write_text(json.dumps(data, ensure_ascii=False, indent=2), encoding="utf-8")


def _decoded_ctx(params: Dict[str, str]) -> Dict[str, str]:
    return {
        "mid": params["mid"],
        "season": unquote_plus(params["season"]),
        "date": unquote_plus(params["date"]),
        "division": unquote_plus(params["division"]),
        "j1": unquote_plus(params["j1"]),
        "j2": unquote_plus(params["j2"]),
        "sig": params.get("sig", ""),
    }


# =================
# Router + helpers
# =================
router = APIRouter()


def build_submit_link(
    *, base_url: str, mid: str, season: str, date: str, division: str, j1: str, j2: str
) -> str:
    params = {"mid": mid, "season": season, "date": date, "division": division, "j1": j1, "j2": j2}
    sig = _make_sig(params)
    query = _canonical_query(params) + f"&sig={sig}"
    return f"{base_url.rstrip('/')}/submit?{query}"


def _norm_qs(request: Request) -> Dict[str, str]:
    # Compat con /submit-result (legacy)
    q = request.query_params
    return {
        "mid": q.get("mid") or q.get("id") or "",
        "season": q.get("season") or "",
        "date": q.get("date") or q.get("fecha") or "",
        "division": q.get("division") or q.get("div") or "",
        "j1": q.get("j1") or "",
        "j2": q.get("j2") or "",
        "sig": q.get("sig") or "",
    }


# ==========  /submit  ==========
@router.get("/submit", response_class=HTMLResponse)
def submit_get(
    request: Request,
    mid: str,
    season: str,
    date: str,
    division: str,
    j1: str,
    j2: str,
    sig: str = "",
):
    params = {"mid": mid, "season": season, "date": date, "division": division, "j1": j1, "j2": j2}
    if sig:
        _verify_sig(params, sig)
    return TEMPLATES.TemplateResponse("submit_result.html", {"request": request, **_decoded_ctx({**params, "sig": sig})})


@router.post("/submit", response_class=HTMLResponse)
async def submit_post(
    request: Request,
    mid: str = Form(...),
    season: str = Form(...),
    date: str = Form(...),
    division: str = Form(...),
    j1: str = Form(...),
    j2: str = Form(...),
    sig: str = Form(""),
    s1_j1: str = Form(""),
    s1_j2: str = Form(""),
    s2_j1: str = Form(""),
    s2_j2: str = Form(""),
    s3_j1: str = Form(""),
    s3_j2: str = Form(""),
    s4_j1: str = Form(""),
    s4_j2: str = Form(""),
    s5_j1: str = Form(""),
    s5_j2: str = Form(""),
    submitter: str = Form(""),
):
    params = {"mid": mid, "season": season, "date": date, "division": division, "j1": j1, "j2": j2}
    if sig:
        _verify_sig(params, sig)

    record = {
        **params,
        "sets": {
            "s1": [s1_j1, s1_j2],
            "s2": [s2_j1, s2_j2],
            "s3": [s3_j1, s3_j2],
            "s4": [s4_j1, s4_j2],
            "s5": [s5_j1, s5_j2],
        },
        "submitter": submitter,
    }

    existing = _load_json(PENDING_FILE)
    existing.append(record)
    _save_json(PENDING_FILE, existing)

    return TEMPLATES.TemplateResponse("submit_result_done.html", {"request": request, **_decoded_ctx({**params, "sig": sig})})


# ==========  /submit-result (legacy) ==========
@router.get("/submit-result", response_class=HTMLResponse)
def submit_result_get_legacy(request: Request):
    params = _norm_qs(request)
    sig = params.pop("sig", "")
    if sig:
        _verify_sig(params, sig)
    return TEMPLATES.TemplateResponse("submit_result.html", {"request": request, **_decoded_ctx({**params, "sig": sig})})


@router.post("/submit-result", response_class=HTMLResponse)
async def submit_result_post_legacy(request: Request):
    form = await request.form()
    mid = form.get("mid") or form.get("id") or ""
    season = form.get("season") or ""
    date = form.get("date") or form.get("fecha") or ""
    division = form.get("division") or form.get("div") or ""
    j1 = form.get("j1") or ""
    j2 = form.get("j2") or ""
    sig = form.get("sig", "")

    def g(k: str) -> str:
        return form.get(k, "")

    params = {"mid": mid, "season": season, "date": date, "division": division, "j1": j1, "j2": j2}
    if sig:
        _verify_sig(params, sig)

    record = {
        **params,
        "sets": {
            "s1": [g("s1_j1"), g("s1_j2")],
            "s2": [g("s2_j1"), g("s2_j2")],
            "s3": [g("s3_j1"), g("s3_j2")],
            "s4": [g("s4_j1"), g("s4_j2")],
            "s5": [g("s5_j1"), g("s5_j2")],
        },
        "submitter": form.get("submitter", ""),
    }

    existing = _load_json(PENDING_FILE)
    existing.append(record)
    _save_json(PENDING_FILE, existing)

    return TEMPLATES.TemplateResponse("submit_result_done.html", {"request": request, **_decoded_ctx({**params, "sig": sig})})


# ==========================
# ====== ADMIN ROUTES ======
# ==========================
def _require_admin(request: Request) -> None:
    token = request.query_params.get("token") or request.headers.get("x-admin-token") or ""
    if token != ADMIN_TOKEN:
        raise HTTPException(status_code=401, detail="Token inválido")


def _flatten_item(it: Dict[str, Any]) -> Dict[str, Any]:
    out = {
        "id": it.get("mid", ""),
        "fecha": it.get("date", ""),
        "division": it.get("division", ""),
        "j1": it.get("j1", ""),
        "j2": it.get("j2", ""),
    }
    sets = it.get("sets") or {}
    for i in range(1, 6):
        a, b = (sets.get(f"s{i}", ["", ""]) + ["", ""])[:2]
        out[f"s{i}_j1"] = a
        out[f"s{i}_j2"] = b
    return out


@router.get("/admin/review", response_class=HTMLResponse)
def admin_review(request: Request):
    _require_admin(request)
    pending = _load_json(PENDING_FILE)
    items = [_flatten_item(x) for x in pending]
    token = request.query_params.get("token", "")
    return TEMPLATES.TemplateResponse("admin_review.html", {"request": request, "items": items, "token": token})


@router.get("/admin/reject")
def admin_reject(request: Request, id: str):
    _require_admin(request)
    pending = _load_json(PENDING_FILE)
    pending = [x for x in pending if x.get("mid") != id]
    _save_json(PENDING_FILE, pending)
    token = request.query_params.get("token", "")
    return RedirectResponse(url=f"/admin/review?token={quote_plus(token)}", status_code=302)


# ==========================
# ===== APROBAR =====
# ==========================
def _season_to_results_path(season_str: str) -> Path:
    m = re.search(r"(\d+)", season_str or "")
    if not m:
        raise HTTPException(status_code=400, detail="Temporada inválida")
    n = m.group(1)
    return ROOT_DIR / f"RESULTADOS T{n}.xlsx"


def _cell_str(v: Any) -> str:
    from datetime import datetime, date
    if isinstance(v, datetime):
        return v.date().isoformat()
    if isinstance(v, date):
        return v.isoformat()
    return str(v or "").strip()


def _norm_txt(s: str) -> str:
    """minúsculas, sin diacríticos, espacios colapsados"""
    s = str(s or "")
    s = unicodedata.normalize("NFKD", s)
    s = "".join(c for c in s if not unicodedata.combining(c))
    s = s.lower()
    s = re.sub(r"\s+", " ", s)
    return s.strip()


def _norm_date(s: str) -> str:
    """A iso (YYYY-MM-DD) si viene con hora/variantes."""
    from datetime import datetime
    s = str(s or "").strip()
    for fmt in ("%Y-%m-%d", "%Y/%m/%d", "%d/%m/%Y", "%Y-%m-%d %H:%M:%S"):
        try:
            return datetime.strptime(s, fmt).date().isoformat()
        except Exception:
            pass
    return s[:10]


def _looks_like_time(v: Any) -> bool:
    """Heurística para detectar columna de hora en RESULTADOS T6."""
    from datetime import time, datetime
    import pandas as pd

    if v is None:
        return False
    if isinstance(v, time):
        return True
    if isinstance(v, datetime):
        return True
    # strings tipo "09:00" / "09:00:00"
    s = str(v).strip()
    if not s:
        return False
    if ":" in s:
        p = s.split(":")
        if len(p) >= 2 and p[0].isdigit() and p[1].isdigit():
            return True
    # timestamps pandas (por si acaso)
    if "Timestamp" in type(v).__name__:
        return True
    return False


def _infer_results_layout_openpyxl(ws) -> dict:
    """
    Detecta si la hoja tiene columna HORA en B.

    Legacy:
      A fecha, B division, C j1, D j2, E.. sets

    T6:
      A fecha, B hora, C division, D j1, E j2, F.. sets
    """
    # miramos unas cuantas filas para encontrar un ejemplo en col B
    has_time = False
    max_check = min(ws.max_row, 20)
    for r in range(1, max_check + 1):
        v = ws.cell(row=r, column=2).value  # B
        if _looks_like_time(v):
            has_time = True
            break

    if has_time:
        return {"FECHA": 1, "HORA": 2, "DIV": 3, "J1": 4, "J2": 5, "SETS_START": 6}
    return {"FECHA": 1, "HORA": None, "DIV": 2, "J1": 3, "J2": 4, "SETS_START": 5}


def _apply_to_excel(record: Dict[str, Any]) -> None:
    """
    Escribe los sets en la fila correspondiente del Excel de resultados.

    Soporta:
    - formato legacy (T1..T5): A fecha, B división, C J1, D J2, E.. sets
    - formato T6 (con hora):   A fecha, B hora,     C división, D J1, E J2, F.. sets

    IMPORTANTE:
    - Para localizar la fila NO exigimos coincidencia por hora (no viene en el submit).
    - Sí exigimos fecha + división + jugadores.
    """
    from openpyxl import load_workbook

    xlsx = _season_to_results_path(record["season"])
    if not xlsx.exists():
        raise HTTPException(status_code=400, detail=f"No existe el Excel de resultados: {xlsx.name}")

    wb = load_workbook(filename=str(xlsx))
    try:
        ws = wb.worksheets[0]  # siempre 1 hoja

        layout = _infer_results_layout_openpyxl(ws)

        # normalizamos lo que llega del submit
        date_str = _norm_date(record["date"])
        div_str = _norm_txt(record["division"])
        j1_str = _norm_txt(record["j1"])
        j2_str = _norm_txt(record["j2"])

        target_row = None

        # Recorremos TODO (no hay cabeceras)
        for r in range(1, ws.max_row + 1):
            vfecha = _cell_str(ws.cell(row=r, column=layout["FECHA"]).value)
            vdiv = _cell_str(ws.cell(row=r, column=layout["DIV"]).value)
            vj1 = _cell_str(ws.cell(row=r, column=layout["J1"]).value)
            vj2 = _cell_str(ws.cell(row=r, column=layout["J2"]).value)

            if not vj1 and not vj2:
                continue

            if (
                _norm_date(vfecha) == date_str
                and _norm_txt(vdiv) == div_str
                and _norm_txt(vj1) == j1_str
                and _norm_txt(vj2) == j2_str
            ):
                target_row = r
                break

        if not target_row:
            raise HTTPException(status_code=400, detail="No encuentro el partido en el Excel (fecha/división/jugadores)")

        # Escribir sets
        first_set_col = layout["SETS_START"]
        for i in range(1, 6):
            a, b = (record["sets"].get(f"s{i}", ["", ""]) + ["", ""])[:2]
            c1 = first_set_col + (i - 1) * 2
            c2 = c1 + 1
            ws.cell(row=target_row, column=c1).value = (a or None)
            ws.cell(row=target_row, column=c2).value = (b or None)

        wb.save(str(xlsx))
    finally:
        wb.close()


@router.get("/admin/approve")
def admin_approve(request: Request, id: str):
    _require_admin(request)

    pending = _load_json(PENDING_FILE)
    keep: List[Dict[str, Any]] = []
    moved: List[Dict[str, Any]] = []

    for x in pending:
        if x.get("mid") == id:
            moved.append(x)
        else:
            keep.append(x)

    if moved:
        try:
            _apply_to_excel(moved[-1])
        except HTTPException:
            raise
        except Exception as e:
            raise HTTPException(status_code=500, detail=f"Error escribiendo en Excel: {e}")

        approved = _load_json(APPROVED_FILE)
        approved.extend(moved)
        _save_json(APPROVED_FILE, approved)

    _save_json(PENDING_FILE, keep)

    token = request.query_params.get("token", "")
    return RedirectResponse(url=f"/admin/review?token={quote_plus(token)}", status_code=302)
